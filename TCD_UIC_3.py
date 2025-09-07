import os
import re
import glob
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# =========================
# CONFIG UTILISATEUR
# =========================
ROOT_PATHS = [
    r"C:\Program Files (x86)\AVEVA",  # ROOT 1
    r"D:\E3D.2.1",                    # ROOT 2
]

# Fichier Excel final (un seul artefact de sortie)
OUTPUT_FOLDER = r"D:\BUREAU-BUREAU-BUREAU-BUREAU-BUREAU\FORMATION E3D ADMIN\ETUDE AVEVA UIC ETC\UIC XML TRANSFORMATION\3-xlsx_tcd_output"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_FOLDER, "TCD_Synthese_UIC_3.xlsx")

# =========================
# ETAPE 0 : Fonctions utilitaires
# =========================
def safe_read_text(path):
    for enc in ("utf-8", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    with open(path, "rb") as fb:
        raw = fb.read()
    return raw.decode("utf-8", errors="replace")

def detect_balise_type(line: str) -> str:
    line = line.strip()
    if line.startswith("<?xml"):
        return "Declaration XML"
    elif line.startswith("</"):
        return "Balise fermante"
    elif "/>" in line:
        return "Balise auto-fermante"
    elif "=" in line and "<" in line and ">" in line:
        return "Balise ouvrante avec attributs"
    elif line.startswith("<") and ">" in line:
        return "Balise ouvrante"
    return "Texte ou inconnu"

def extract_balise_name(line: str) -> str:
    match = re.match(r"<\??/?([a-zA-Z0-9:_-]+)", line)
    return match.group(1) if match else ""

def extract_content(line: str) -> str:
    match_content = re.match(r"<.*?>(.*?)</.*?>", line)
    if match_content:
        return match_content.group(1).strip()
    matches = re.findall(r'([a-zA-Z0-9:_-]+)="(.*?)"', line)
    if matches:
        return ", ".join(f"{key}={value}" for key, value in matches)
    return ""

def extract_extended_content(content: str):
    extended_data = []
    if "=" in content and not re.search(r'=\s*"', content):
        parts = [part.split("=", 1)[1] for part in content.split(", ") if "=" in part]
        extended_data = parts if parts else [""]
    return extended_data

def count_leading_spaces(line: str) -> int:
    return len(line) - len(line.lstrip())

# =========================
# ETAPE 0bis : Index XML -> UIC (par ROOT_PATH)
# =========================
# On repère les références aux .uic dans les .xml : Path="xxx.uic"
UIC_IN_XML_RE = re.compile(r'Path\s*=\s*"([^"]+\.uic)"', re.IGNORECASE)

def build_xml_uic_index(root_path: str) -> dict:
    """
    Retourne un dict: { 'nom_uic.ext' (lower) : set([xml_basename1, xml_basename2, ...]) }
    """
    mapping = {}
    xml_files = glob.glob(os.path.join(root_path, "**", "*.xml"), recursive=True)
    for xf in xml_files:
        try:
            txt = safe_read_text(xf)
        except Exception:
            continue
        hits = set(os.path.basename(m).lower() for m in UIC_IN_XML_RE.findall(txt))
        if not hits:
            continue
        xml_name = os.path.basename(xf)
        for u in hits:
            mapping.setdefault(u, set()).add(xml_name)
    return mapping

# =========================
# ETAPE 0ter : Lecture de tous les .pmlcmd des 2 roots
# =========================
def read_all_pmlcmds(root_paths):
    """
    Retourne une liste [(fullpath, basename, dirpath, content_lower), ...]
    """
    files = []
    for rp in root_paths:
        files.extend(glob.glob(os.path.join(rp, "**", "*.pmlcmd"), recursive=True))
    files = sorted(set(files))
    out = []
    for fp in tqdm(files, desc="Indexation PMLCMD"):
        try:
            text = safe_read_text(fp).lower()
        except Exception:
            continue
        out.append((fp, os.path.basename(fp), os.path.dirname(fp), text))
    return out

# =========================
# ETAPE 1 : Lecture et extraction UIC -> DataFrame (mémoire)
# =========================
def parse_uic_to_dataframe(file_path: str) -> pd.DataFrame:
    # lecture texte (avec fallback encodage)
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            raw_lines = f.readlines()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            raw_lines = f.readlines()

    lines_no_left_space = [ln.lstrip() for ln in raw_lines]

    type_balise_col = [detect_balise_type(ln) for ln in lines_no_left_space]
    balise_col = [extract_balise_name(ln) for ln in lines_no_left_space]
    contenu_col = [extract_content(ln) for ln in lines_no_left_space]

    espaces_col = [count_leading_spaces(ln) for ln in raw_lines]

    # calcul des rangs "Rang 0..20" (paliers 0..40 par pas de 2)
    space_levels = list(range(0, 42, 2))  # 0,2,...,40
    count_dict = {level: 0 for level in space_levels}

    rangs = []
    for spaces in espaces_col:
        rang_values = [0] * 21
        for i, level in enumerate(space_levels):
            if spaces == level:
                count_dict[level] += 1
                rang_values[i] = count_dict[level]
                break
            elif spaces > level:
                rang_values[i] = count_dict[level]
        rangs.append(rang_values)

    code_unique_col = [".".join(map(str, [i + 1] + rangs[i])) for i in range(len(raw_lines))]

    filename = os.path.basename(file_path)
    path_to_file = os.path.dirname(file_path)

    contenu_extended = [extract_extended_content(c) for c in contenu_col]
    max_ext_cols = max((len(ext) for ext in contenu_extended), default=0)

    df = pd.DataFrame({
        "Filename": [filename] * len(raw_lines),
        "PathToFile": [path_to_file] * len(raw_lines),
        "Scripte initial": raw_lines,
        "Index": range(1, len(raw_lines) + 1),
        "Nombre d'espace": espaces_col,
        **{f"Rang {i}": [row[i] for row in rangs] for i in range(21)},
        "Code unique": code_unique_col,
        "Scripte initial sans espace": lines_no_left_space,
        "Type de balise": type_balise_col,
        "Balise": balise_col,
        "Contenu": contenu_col
    })

    # colonnes Contenu_extended_*
    for i in range(max_ext_cols):
        df[f"Contenu_extended_{i+1}"] = [
            ext[i] if i < len(ext) else (cont if i == 0 else "")
            for cont, ext in zip(contenu_col, contenu_extended)
        ]

    return df

# =========================
# ETAPE 2 : Enrichissement par Code unique 2 (mémoire)
# =========================
def enrich_dataframe_for_uic(df: pd.DataFrame) -> pd.DataFrame:
    required = ["Rang 0", "Rang 1", "Rang 2", "Nombre d'espace", "Balise", "Contenu_extended_1"]
    if not all(col in df.columns for col in required):
        for col in required:
            if col not in df.columns:
                df[col] = ""

    df["Code unique 2"] = df[["Rang 0", "Rang 1", "Rang 2"]].astype(str).agg(".".join, axis=1)

    rang_mapping = {i: f"Rang {i//2}" for i in range(0, 40, 2)}
    df["Rang selon Nombre d'espace"] = df["Nombre d'espace"].map(rang_mapping).fillna(
        df["Nombre d'espace"].apply(lambda x: f"Rang {int(x)//2 if pd.notna(x) else 0}")
    )

    grouped = df.groupby("Code unique 2").agg(list).to_dict(orient="index")

    def get_balise_rang2(code_unique):
        if code_unique in grouped:
            try:
                for rang, bal in zip(grouped[code_unique]["Rang selon Nombre d'espace"], grouped[code_unique]["Balise"]):
                    if rang == "Rang 2":
                        return bal
            except Exception:
                pass
        return ""

    def get_contenu_extended1(code_unique, target_balise):
        if code_unique in grouped:
            try:
                for bal, cont in zip(grouped[code_unique]["Balise"], grouped[code_unique]["Contenu_extended_1"]):
                    if bal == target_balise:
                        return cont
            except Exception:
                pass
        return ""

    df["Balise de Rang 2 selon Code Unique 2"] = df["Code unique 2"].apply(get_balise_rang2)
    df["Contenu_extended_1 de Balise de Rang 2 selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, get_balise_rang2(x))
    )
    df["Contenu_extended_1 de Caption selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "Caption")
    )
    df["Contenu_extended_1 de Type selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "Type")
    )
    df["Contenu_extended_1 de Key selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "Key")
    )
    df["Contenu_extended_1 de Category selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "Category")
    )
    df["Contenu_extended_1 de CommandBarDisplayStyle selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "CommandBarDisplayStyle")
    )
    df["Contenu_extended_1 de FormKey selon Code Unique 2"] = df["Code unique 2"].apply(
        lambda x: get_contenu_extended1(x, "FormKey")
    )

    def get_namespace_value(local_df: pd.DataFrame):
        if "Balise" in local_df.columns and "Contenu_extended_1" in local_df.columns:
            row = local_df[local_df["Balise"] == "Namespace"]
            if not row.empty:
                return row["Contenu_extended_1"].values[0]
        return ""

    df["Namespace_of_Filename"] = get_namespace_value(df)
    return df

# =========================
# ETAPE 3 : Fusion / colonnes / écriture
# =========================
# Colonnes avec B et C déjà insérées, puis S/T en fin
COLUMNS_ORDER = [
    "Code unique 2",
    "XML (Path 1)",                       # B
    "XML (Path 2)",                       # C
    "Filename",
    "Namespace_of_Filename",
    "PathToFile",
    "Contenu_extended_1 de Caption selon Code Unique 2",
    "Balise de Rang 2 selon Code Unique 2",
    "Contenu_extended_1 de Category selon Code Unique 2",
    "Contenu_extended_1 de CommandBarDisplayStyle selon Code Unique 2",
    "Contenu_extended_1 de Balise de Rang 2 selon Code Unique 2",  # -> Name (col K)
    "Contenu_extended_1 de Type selon Code Unique 2",
    "Contenu_extended_1 de Key selon Code Unique 2",               # -> Key (col M)
    "Contenu_extended_1 de FormKey selon Code Unique 2",
    "Balise", "Contenu_extended_1",
    "Rang selon Nombre d'espace", "Type de balise",
    "PMLCMD (files)",                    # S
    "PMLCMD (paths)",                    # T
]

COLUMN_RENAME = {
    "Code unique 2": "Code unique 2",
    "XML (Path 1)": "XML (Path 1)",
    "XML (Path 2)": "XML (Path 2)",
    "Filename": "Filename",
    "Namespace_of_Filename": "Namespace_of_Filename",
    "PathToFile": "PathToFile",
    "Contenu_extended_1 de Caption selon Code Unique 2": "Tool (Caption)",
    "Balise de Rang 2 selon Code Unique 2": "Control Type",
    "Contenu_extended_1 de Category selon Code Unique 2": "Category",
    "Contenu_extended_1 de CommandBarDisplayStyle selon Code Unique 2": "CommandBarDisplayStyle",
    "Contenu_extended_1 de Balise de Rang 2 selon Code Unique 2": "Name",
    "Contenu_extended_1 de Type selon Code Unique 2": "Type",
    "Contenu_extended_1 de Key selon Code Unique 2": "Key (Command) !this.key",
    "Contenu_extended_1 de FormKey selon Code Unique 2": "FormKey",
    "Balise": "Balise",
    "Contenu_extended_1": "ApplicationContext",
    "Rang selon Nombre d'espace": "Rang",
    "Type de balise": "Type de balise",
    "PMLCMD (files)": "PMLCMD (files)",
    "PMLCMD (paths)": "PMLCMD (paths)",
}

def write_synthesis_excel(df_final: pd.DataFrame, output_file: str):
    df_final = df_final.copy()
    df_final["ColorGroup"] = (df_final["Code unique 2"] != df_final["Code unique 2"].shift()).cumsum()

    wb = Workbook()
    ws = wb.active
    ws.append(COLUMNS_ORDER)

    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col_idx in range(1, len(COLUMNS_ORDER) + 1):
        ws.cell(row=1, column=col_idx).fill = fill_yellow

    col_widths = {i: len(COLUMNS_ORDER[i - 1]) for i in range(1, len(COLUMNS_ORDER) + 1)}

    previous_code = None
    current_fill = fill_white

    for idx, row in df_final.iterrows():
        row_code = row["Code unique 2"]
        if row_code != previous_code:
            current_fill = fill_white if current_fill == fill_blue else fill_blue

        out_row = [row.get(col, "") for col in COLUMNS_ORDER]
        ws.append(out_row)

        for c_idx, value in enumerate(out_row, start=1):
            cell = ws.cell(row=idx + 2, column=c_idx)
            cell.fill = current_fill
            vlen = len(str(value)) if value is not None else 0
            col_widths[c_idx] = max(col_widths[c_idx], vlen)

        previous_code = row_code

    for c_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(width + 2, 50)

    for col_idx, original in enumerate(COLUMNS_ORDER, start=1):
        if original in COLUMN_RENAME:
            ws.cell(row=1, column=col_idx).value = COLUMN_RENAME[original]

    ws.auto_filter.ref = ws.dimensions

    last_col_idx = ws.max_column
    last_col_letter = get_column_letter(last_col_idx)
    ws[f"{last_col_letter}1"] = "Index_of_Code unique 2"
    ws[f"{last_col_letter}1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(output_file)

# =========================
# PIPELINE GLOBAL
# =========================
def main():
    # 0) Index XML -> UIC (path1 / path2)
    print("Indexation des XML (Path 1)...")
    xml_index_1 = build_xml_uic_index(ROOT_PATHS[0])
    print(f"  {len(xml_index_1)} clés UIC référencées dans des XML du Path 1")

    print("Indexation des XML (Path 2)...")
    xml_index_2 = build_xml_uic_index(ROOT_PATHS[1])
    print(f"  {len(xml_index_2)} clés UIC référencées dans des XML du Path 2")

    # 0bis) Charger tous les .pmlcmd (contenu en minuscules)
    pml_list = read_all_pmlcmds(ROOT_PATHS)  # [(full, base, dir, text_lower), ...]
    print(f"{len(pml_list)} fichiers PMLCMD indexés.")

    # 1) Lister tous les .uic sous les deux root paths
    uic_files = []
    for rp in ROOT_PATHS:
        uic_files.extend(glob.glob(os.path.join(rp, "**", "*.uic"), recursive=True))
    uic_files = sorted(set(uic_files))

    if not uic_files:
        print("Aucun fichier UIC trouvé dans les répertoires fournis.")
        return

    print(f"{len(uic_files)} fichiers UIC détectés.")
    all_enriched = []

    # 2) Pour chaque UIC : extraction -> enrichissement -> collecte (mémoire)
    for file_path in tqdm(uic_files, desc="Traitement UIC (sans intermédiaires)"):
        try:
            df_raw = parse_uic_to_dataframe(file_path)
            df_enriched = enrich_dataframe_for_uic(df_raw)

            # Ajout des 2 colonnes XML (Path 1 / Path 2) constantes par fichier
            uic_key = os.path.basename(file_path).lower()
            xmls_1 = " | ".join(sorted(xml_index_1.get(uic_key, [])))
            xmls_2 = " | ".join(sorted(xml_index_2.get(uic_key, [])))
            df_enriched["XML (Path 1)"] = xmls_1
            df_enriched["XML (Path 2)"] = xmls_2

            # S'assurer des colonnes de sortie
            for col in COLUMNS_ORDER:
                if col not in df_enriched.columns:
                    df_enriched[col] = ""
            df_enriched = df_enriched[COLUMNS_ORDER]

            all_enriched.append(df_enriched)
        except Exception as e:
            print(f"[AVERTISSEMENT] Échec fichier: {file_path} -> {e}")

    if not all_enriched:
        print("Aucune donnée exploitable.")
        return

    # 3) Fusion mémoire
    df_final = pd.concat(all_enriched, ignore_index=True)

    # 3bis) REMPLISSAGE des colonnes S/T à partir des .pmlcmd
    # tokens issus des colonnes 'Name' (col K avant renommage) et 'Key (Command)' (col M avant renommage)
    COL_NAME = "Contenu_extended_1 de Balise de Rang 2 selon Code Unique 2"
    COL_KEY  = "Contenu_extended_1 de Key selon Code Unique 2"

    # Set de tokens en minuscule
    tokens = set()
    if COL_NAME in df_final.columns:
        tokens |= set(str(x).strip().lower() for x in df_final[COL_NAME].dropna().astype(str) if str(x).strip() and str(x).lower() != "nan")
    if COL_KEY in df_final.columns:
        tokens |= set(str(x).strip().lower() for x in df_final[COL_KEY].dropna().astype(str) if str(x).strip() and str(x).lower() != "nan")

    # Construire un index token -> [(file_base, dir_path), ...]
    from collections import defaultdict
    token_to_files = defaultdict(list)

    if tokens and pml_list:
        print("Recherche des correspondances dans les PMLCMD…")
        for tok in tqdm(tokens, desc="Matching tokens"):
            for full, base, dpath, txt in pml_list:
                if tok and tok in txt:  # recherche sous-chaîne, insensible à la casse
                    token_to_files[tok].append((base, dpath))

    def find_pml_matches(name_val, key_val):
        files = []
        paths = []
        for tok in (str(name_val).strip().lower(), str(key_val).strip().lower()):
            if tok and tok != "nan":
                for base, dpath in token_to_files.get(tok, []):
                    files.append(base); paths.append(dpath)
        # déduplication en conservant l'ordre
        seen = set()
        f2, p2 = [], []
        for b, p in zip(files, paths):
            k = (b, p)
            if k not in seen:
                seen.add(k)
                f2.append(b); p2.append(p)
        return " | ".join(f2), " | ".join(p2)

    if COL_NAME in df_final.columns and COL_KEY in df_final.columns:
        pairs = df_final.apply(lambda r: find_pml_matches(r[COL_NAME], r[COL_KEY]), axis=1)
        df_final["PMLCMD (files)"] = [p[0] for p in pairs]
        df_final["PMLCMD (paths)"] = [p[1] for p in pairs]
    else:
        df_final["PMLCMD (files)"] = ""
        df_final["PMLCMD (paths)"] = ""

    # 4) Ecriture Excel final (unique)
    write_synthesis_excel(df_final, OUTPUT_FILE)
    print(f"Fichier généré : {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
