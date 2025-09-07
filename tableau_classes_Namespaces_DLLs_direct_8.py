import os
import re
import subprocess
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
ilspy_path = r"C:\Users\Nicolas JF Martin\.dotnet\tools\ilspycmd.exe"
dll_roots = [
    r"C:\Program Files (x86)\AVEVA\Everything3D2.10",
    r"D:\E3D.2.1",
    r"C:\Users\Public\Documents\AVEVA"
]
decompile_root = r"D:\BUREAU-BUREAU-BUREAU-BUREAU-BUREAU\FORMATION E3D ADMIN\DLL decompilation\transposition_dll_all"
global_excel_path = r"D:\BUREAU-BUREAU-BUREAU-BUREAU-BUREAU\FORMATION E3D ADMIN\ETUDE AVEVA UIC ETC\code_unique_dll_static_ILSpy.xlsx"

os.makedirs(decompile_root, exist_ok=True)

# === COLLECTE DES DLL ===
dll_paths = []
for root in dll_roots:
    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            if filename.lower().endswith(".dll"):
                dll_paths.append(os.path.join(dirpath, filename))

all_rows = []
for dll_path in tqdm(dll_paths, desc="Analyse des DLL"):
    dll_name = os.path.basename(dll_path)
    dll_short_name = os.path.splitext(dll_name)[0]
    dll_output_dir = os.path.join(decompile_root, dll_short_name)
    os.makedirs(dll_output_dir, exist_ok=True)

    cmd = [ilspy_path, dll_path, "-p", "-o", dll_output_dir]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"❌ Échec décompilation {dll_name}")
        continue

    version = ""
    for root_dir, _, files in os.walk(dll_output_dir):
        for file in files:
            if "AssemblyInfo" in file and file.endswith(".cs"):
                try:
                    with open(os.path.join(root_dir, file), "r", encoding="utf-8", errors="ignore") as f:
                        for line in f:
                            if "AssemblyVersion" in line:
                                version = line.strip().split("\"")[1]
                                break
                except:
                    pass
        if version:
            break

    per_dll_rows = []
    for root_dir, _, files in os.walk(dll_output_dir):
        for file in files:
            if file.endswith(".cs"):
                namespace = ""
                pmlnet_entries = []
                try:
                    with open(os.path.join(root_dir, file), "r", encoding="utf-8", errors="ignore") as f:
                        lines = f.readlines()

                    for line in lines:
                        if line.strip().startswith("namespace"):
                            match = re.search(r'namespace\s+([\w\.]+)', line)
                            if match:
                                namespace = match.group(1)
                            break

                    i = 0
                    while i < len(lines):
                        line = lines[i]
                        if '[PMLNetCallable]' in line:
                            signature_lines = []
                            original_line_num = i + 1
                            i += 1
                            while i < len(lines):
                                current_line = lines[i].strip()
                                if current_line == '':
                                    i += 1
                                    continue
                                signature_lines.append(current_line)
                                if '(' in ''.join(signature_lines) and ')' in ''.join(signature_lines):
                                    break
                                elif '(' not in ''.join(signature_lines):
                                    break
                                i += 1
                            signature = ' '.join(signature_lines)
                            pmlnet_entries.append(('[PMLNetCallable]', signature, original_line_num))
                        i += 1

                    if not pmlnet_entries:
                        pmlnet_entries.append(("", "", ""))

                    for annotation, full_signature, line_number in pmlnet_entries:
                        concat = f"{dll_short_name}/{namespace}/{file}"
                        concat2 = f"{concat}/{full_signature}"
                        row = [
                            dll_path,
                            version,
                            dll_short_name,
                            annotation,
                            full_signature,
                            line_number,
                            file,
                            namespace,
                            concat,
                            concat2
                        ]
                        per_dll_rows.append(row)
                        all_rows.append(row)

                except:
                    continue

    if per_dll_rows:
        per_dll_df = pd.DataFrame(per_dll_rows, columns=[
            "Chemin complet DLL", "Version Assembly", "Nom DLL", "[PMLNetCallable]", "PMLNetCallable", "Numéro de ligne",
            "Fichier .cs", "Namespace", "Concatener", "Concatener 2"])
        dll_excel_path = os.path.join(dll_output_dir, f"{dll_short_name}.xlsx")
        per_dll_df.to_excel(dll_excel_path, index=False)

# === GLOBAL DATAFRAME ===
df = pd.DataFrame(all_rows, columns=[
    "Chemin complet DLL", "Version Assembly", "Nom DLL", "[PMLNetCallable]", "PMLNetCallable", "Numéro de ligne",
    "Fichier .cs", "Namespace", "Concatener", "Concatener 2"])

# CODES UNIQUES
for i, (col, prefix, digits) in enumerate([
    ("Nom DLL", "DLL", 5),
    ("Namespace", "NS", 4),
    ("Fichier .cs", "CS", 4),
    ("PMLNetCallable", "CALL", 4)
]):
    code_col = f"Code unique - Niveau {i+1}"
    df[code_col] = pd.factorize(df[col])[0] + 1
    df[code_col] = df[code_col].apply(lambda x: f"{prefix}{x:0{digits}d}")

# IDENTIFIANTS INTÉGRAUX
df["Code unique intégral"] = df["Code unique - Niveau 1"] + "-" + df["Code unique - Niveau 2"] + "-" + df["Code unique - Niveau 3"]
df["Code unique intégral 2"] = df["Code unique intégral"] + "-" + df["Code unique - Niveau 4"]

# === ANALYSE C# TYPE ET NOM ===
CSHARP_FAMILIES = {
    'class': 'Class', 'interface': 'Interface', 'struct': 'Struct', 'enum': 'Enum', 'delegate': 'Delegate',
    'event': 'Event', 'void': 'Method', 'string': 'Property', 'int': 'Property', 'double': 'Property',
    'float': 'Property', 'bool': 'Property', 'Hashtable': 'Property', 'DbElement': 'Property', 'BasicPoint': 'Property'
}

def detect_family(line):
    line = line.strip()

    # Étape 0 : ligne vide ou seulement modificateur
    if re.fullmatch(r'\b(public|private|internal|protected)\b', line):
        return "INCOMPLETE"

    # Étape 1 : détection standard avec types connus
    for keyword, family in CSHARP_FAMILIES.items():
        if re.search(rf'\b(public|private|internal|protected)(\s+(unsafe|sealed|static|new))*\s+{keyword}\b', line):
            return family

    # Étape 2 : classe scellée ou unsafe
    if re.search(r'\b(public|private|internal|protected)\s+(unsafe\s+)?sealed\s+class\b', line):
        return "Class"

    # Étape 3 : méthode avec parenthèses classiques
    if re.search(r'\b(public|private|internal|protected)(\s+(unsafe|sealed|static|new))*\s+\w+\s+\w+\s*\(', line):
        return "Method"

    # Étape 4 : propriété avec `=>` (expression-bodied)
    if re.search(r'\b(public|private|internal|protected)(\s+(unsafe|sealed|static|new))*\s+\w+\s+\w+\s*=>', line):
        return "Property"

    # Étape 5 : propriété simple sans corps
    if re.search(r'\b(public|private|internal|protected)(\s+(unsafe|sealed|static|new))*\s+\w+\s+\w+$', line):
        return "Property"

    # Étape 6 : fallback AVEVA (nom seul suivi de parenthèse)
    if re.search(r'\b(public|private|internal|protected)(\s+unsafe)?\s+\w+\s*\(', line):
        return "Method AVEVA"

    return "UNKNOWN"

def extract_name(line, family):
    try:
        if family == "Class":
            return re.findall(r'class\s+(\w+)', line)[0]
        elif family == "Interface":
            return re.findall(r'interface\s+(\w+)', line)[0]
        elif family == "Struct":
            return re.findall(r'struct\s+(\w+)', line)[0]
        elif family == "Enum":
            return re.findall(r'enum\s+(\w+)', line)[0]
        elif family == "Delegate":
            return re.findall(r'delegate\s+\w+\s+(\w+)', line)[0]
        elif family == "Event":
            return re.findall(r'event\s+\S+\s+(\w+)', line)[0]
        elif family == "Constructor":
            return re.findall(r'new\s+(\w+)', line)[0]
        elif family == "Method":
            return re.findall(r'\b\w+\s+(\w+)\s*\(', line)[0]
        elif family == "Method AVEVA":
            return re.findall(r'\b(public|private|internal|protected)\s+(\w+)\s*\(', line)[0][1]
        elif family == "Property":
            match = re.findall(r'\b\w+\s+(\w+)', line)
            return match[-1] if match else ""
    except:
        return ""
    return ""

# Application sur la colonne contenant les définitions C#
df["Famille C#"] = df["PMLNetCallable"].fillna("").apply(detect_family)
df["Nom élément"] = df.apply(lambda row: extract_name(row["PMLNetCallable"], row["Famille C#"]), axis=1)

# RÉORGANISATION ET EXPORT
final_cols = [
    "Chemin complet DLL", "Version Assembly", "Nom DLL", "[PMLNetCallable]", "PMLNetCallable", "Numéro de ligne",
    "Famille C#", "Nom élément", "Fichier .cs", "Namespace", "Concatener", "Concatener 2",
    "Code unique - Niveau 1", "Code unique - Niveau 2", "Code unique - Niveau 3", "Code unique - Niveau 4",
    "Code unique intégral", "Code unique intégral 2"
]
df = df[final_cols]
df.to_excel(global_excel_path, index=False)

# AUTO FORMAT
wb = load_workbook(global_excel_path)
ws = wb.active
ws.auto_filter.ref = ws.dimensions
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2
wb.save(global_excel_path)
print(f"\n✅ Fichier global enrichi généré avec succès : {global_excel_path}")